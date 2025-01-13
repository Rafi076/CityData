from tkinter import *
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
from tkcalendar import Calendar
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib
import re
from qr_code_generator import generate_qr_code  # Import the QR code generation function

# Create the workbook if it doesn't exist
file = pathlib.Path('User_record.xlsx')


## 58755984418
## Melissa Adams
def submit():
    name = nameValue.get()
    contact = contactValue.get()
    age = AgeValue.get()
    gender = gender_combobox.get()
    address = addressEntry.get(1.0, END)
    email = emailValue.get()
    dob = dobEntry.get()
    url = urlValue.get()
    Disease = diseaseValue.get()
    Blood  = bloodValue.get()
    Hobby = hobbyValue.get()
    Occupation = occupationValue.get()
    Education = EeducationValue.get()
    Marital = maritalValue.get()
    Sleeping = sleepingValue.get()
    Income = incomeValue.get()
    Stress = stressValue.get()
    Savings = savingsValue.get()





    # Validation
    if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
        messagebox.showerror("Error", "Invalid email format")
        return
    if not contact.isdigit() or len(contact) != 11:
        messagebox.showerror("Error", "Invalid phone number. Enter 11 digits.")
        return
    if url and not re.match(r"http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\\(\\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+", url):
        messagebox.showerror("Error", "Invalid URL format.")
        return

    try:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active

        # Delete the previous record if exists
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
            if row[1].value == contact:  # Assuming column 2 is "Contact"
                sheet.delete_rows(row[0].row, 1)
                break

        # Append the new data
        sheet.append([name, contact, age, gender, address.strip(), email, dob, url,Disease,Blood,Hobby,Occupation,Education,Marital,Sleeping,Income,Stress,Savings])
        workbook.save(file)

        messagebox.showinfo("Success", "Data updated successfully!")
        clear()  # Clear the form fields

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")


def clear():
    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    emailValue.set('')
    urlValue.set('')
    addressEntry.delete(1.0, END)
    dobEntry.delete(0, END)

    diseaseValue.set('')
    bloodValue.set('')
    hobbyValue.set('')
    occupationValue.set('')
    EeducationValue.set('')
    maritalValue.set('')
    sleepingValue.set('')
    incomeValue.set('')
    stressValue.set('')
    savingsValue.set('')
    
    

def open_calendar():
    top = Toplevel(root)
    top.title("Select Date")

    # Create the calendar widget
    cal = Calendar(top, selectmode="day", year=2000, month=1, day=1)
    cal.pack(pady=20)

    def grab_date():
        dobEntry.delete(0, END)
        dobEntry.insert(0, cal.get_date())  # Get the selected date and insert into the entry
        top.destroy()  # Close the calendar window

    Button(top, text="Select", command=grab_date).pack(pady=20)




def search():
    contact = searchValue.get()
    if not contact.isdigit() or len(contact) != 11:
        messagebox.showerror("Error", "Invalid phone number. Enter 11 digits.")
        return

    try:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active

        # Iterate through rows to find the matching phone number
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
            if row[1].value == contact:  # Assuming column 2 is "Contact"
                # Populate entry fields with the existing data
                nameValue.set(row[0].value)
                contactValue.set(row[1].value)
                AgeValue.set(row[2].value)
                gender_combobox.set(row[3].value)
                addressEntry.delete(1.0, END)
                addressEntry.insert(1.0, row[4].value)
                emailValue.set(row[5].value)
                dobEntry.delete(0, END)
                dobEntry.insert(0, row[6].value)
                urlValue.set(row[7].value if len(row) > 7 else '')

                diseaseValue.set(row[8].value)
                bloodValue.set(row[9].value)
                hobbyValue.set(row[10].value)
                occupationValue.set(row[11].value)
                EeducationValue.set(row[12].value)
                maritalValue.set(row[13].value)
                sleepingValue.set(row[14].value)
                incomeValue.set(row[15].value)
                stressValue.set(row[16].value)
                savingsValue.set(row[17].value)


                messagebox.showinfo("Success", "Record found. Edit the details and resubmit.")
                workbook.close()
                return

        messagebox.showerror("Error", "Record not found.")
        workbook.close()

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")


def refresh():
    # Clear the search field
    searchValue.set('')

    # Clear the form fields
    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    emailValue.set('')
    urlValue.set('')
    addressEntry.delete(1.0, END)
    dobEntry.delete(0, END)
    gender_combobox.set(' Male')

    diseaseValue.set('')
    bloodValue.set('')
    hobbyValue.set('')
    occupationValue.set('')
    EeducationValue.set('')
    maritalValue.set('')
    sleepingValue.set('')
    incomeValue.set('')
    stressValue.set('')
    savingsValue.set('')

    messagebox.showinfo("Refreshed", "Search field and form reset successfully!")



root = tk.Tk()
root.title("Data Entry")
root.geometry('1200x700+700+100')
root.resizable(False, False)
root.configure(bg="#87CEEB")

# Icon
# icon_image = PhotoImage(file="icon.png")
# root.iconphoto(False, icon_image)

# Heading
Label(root, text="Please fill This form: ", font="arial", bg="#87CEEB", fg="#000000").place(x=20, y=20)

# Labels
Label(root, text='Name: ', font=23, bg="#87CEEB", fg="black").place(x=20, y=80)
Label(root, text='Search by Phone:', font=23, bg="#87CEEB", fg="black").place(x=700, y=80)
Label(root, text='Contact: ', font=23, bg="#87CEEB", fg="black").place(x=20, y=120)
Label(root, text='Age:', font=23, bg="#87CEEB", fg="black").place(x=20, y=160)
Label(root, text='Gender: ', font=23, bg="#87CEEB", fg="black").place(x=315, y=160)
Label(root, text='Address: ', font=23, bg="#87CEEB", fg="black").place(x=20, y=200)
Label(root, text='Email: ', font=23, bg="#87CEEB", fg="black").place(x=20, y=260)
Label(root, text='Date of Birth: ', font=23, bg="#87CEEB", fg="black").place(x=20, y=300)
Label(root, text='Website URL: ', font=23, bg="#87CEEB", fg="black").place(x=20, y=340)  # New URL label

Label(root, text='Disease: ', font=23, bg="#87CEEB", fg="black").place(x=20, y=380)
Label(root, text='Blood: ', font=23, bg="#87CEEB", fg="black").place(x=20, y=420)
Label(root, text='Hobby: ', font=23, bg="#87CEEB", fg="black").place(x=20, y=460)
Label(root, text='Occupation: ', font=23, bg="#87CEEB", fg="black").place(x=315, y=380)
Label(root, text='Education: ', font=23, bg="#87CEEB", fg="black").place(x=315, y=420)
Label(root, text='Marital: ', font=23, bg="#87CEEB", fg="black").place(x=315, y=460)
Label(root, text='Sleeping: ', font=23, bg="#87CEEB", fg="black").place(x=600, y=380)
Label(root, text='Income: ', font=23, bg="#87CEEB", fg="black").place(x=600, y=420)
Label(root, text='Stress: ', font=23, bg="#87CEEB", fg="black").place(x=600, y=460)
Label(root, text='Savings: ', font=23, bg="#87CEEB", fg="black").place(x=870, y=380)

# Entry fields
nameValue = StringVar()
contactValue = StringVar()
AgeValue = StringVar()
emailValue = StringVar()
urlValue = StringVar()  # New URL variable
searchValue = StringVar()
diseaseValue = StringVar()
bloodValue= StringVar()
hobbyValue= StringVar()
occupationValue= StringVar()
EeducationValue= StringVar()
maritalValue= StringVar()
sleepingValue= StringVar()
incomeValue= StringVar()
stressValue= StringVar()
savingsValue= StringVar()


nameEntry = Entry(root, textvariable=nameValue, width=45, bd=2, font=20)
searchEntry = Entry(root, textvariable=searchValue, width=15, bd=2, font=20)
contactEntry = Entry(root, textvariable=contactValue, width=15, bd=2, font=20)
ageEntry = Entry(root, textvariable=AgeValue, width=15, bd=2, font=20)
emailEntry = Entry(root, textvariable=emailValue, width=37, bd=2, font=20)
urlEntry = Entry(root, textvariable=urlValue, width=37, bd=2, font=20)  # New URL entry
addressEntry = Text(root, width=51, height=2, bd=2)
dobEntry = Entry(root, width=15, bd=2, font=20)  # Text field for date of birth
diseaseEntry =  Entry(root, textvariable=diseaseValue, width=8, bd=2, font=20)
bloodEntry=  Entry(root, textvariable=bloodValue, width=8, bd=2, font=20)
hobbyEntry=  Entry(root, textvariable=hobbyValue, width=8, bd=2, font=20)
occupationEntry=  Entry(root, textvariable=occupationValue, width=8, bd=2, font=20)
EeducationEntry=  Entry(root, textvariable=EeducationValue, width=8, bd=2, font=20)
maritalEntry=  Entry(root, textvariable=maritalValue, width=8, bd=2, font=20)
sleepingEntry=  Entry(root, textvariable=sleepingValue, width=8, bd=2, font=20)
incomeEntry=  Entry(root, textvariable=incomeValue, width=8, bd=2, font=20)
stressEntry=  Entry(root, textvariable=stressValue, width=8, bd=2, font=20)
savingsEntry=  Entry(root, textvariable=savingsValue, width=8, bd=2, font=20)

nameEntry.place(x=100, y=80)
searchEntry.place(x=900, y=80)
contactEntry.place(x=110, y=120)
ageEntry.place(x=100, y=160)
addressEntry.place(x=110, y=200)
emailEntry.place(x=100, y=260)
dobEntry.place(x=150, y=300)
urlEntry.place(x=160, y=340)  # Place the URL entry

diseaseEntry.place(x=150  ,y=380 )
bloodEntry.place(x=150  ,y=420 )
hobbyEntry.place(x=150  ,y=460 )
occupationEntry.place(x=430  ,y=380 )
EeducationEntry.place(x=430  ,y=420 )
maritalEntry.place(x=430  ,y=460 )
sleepingEntry.place(x=690  ,y=380 )
incomeEntry.place(x=690  ,y=420 )
stressEntry.place(x=690  ,y=460 )
savingsEntry.place(x=960  ,y=380 )

# Gender
gender_combobox = Combobox(root, values=[' Male', ' Female', ' Other'], font='arial 13', state='r', width=10)
gender_combobox.place(x=400, y=160)
gender_combobox.set(' Male')

# Bind date of birth field to open the calendar when clicked
dobEntry.bind("<1>", lambda e: open_calendar())

# Search field and button
Button(root, text="Search", bg="#87CEEB", fg="black", width=10, command=search).place(x=900, y=120)
Button(root, text="Refresh", bg="#87CEEB", fg="black", width=10, command=refresh).place(x=1000, y=120)


# Submit and Clear buttons
Button(root, text="Submit", bg="#87CEEB", fg="black", width=15, height=2, command=submit).place(x=350, y=550)
Button(root, text="Clear", bg="#87CEEB", fg="black", width=15, height=2, command=clear).place(x=500, y=550)
Button(root, text="Exit", bg="#87CEEB", fg="black", width=15, height=2, command=lambda: root.destroy()).place(x=650, y=550)

root.mainloop()



